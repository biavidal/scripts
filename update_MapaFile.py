import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

# Carregar os DataFrames
df_master = pd.read_excel('mapaMaster.xlsx')
df_update = pd.read_excel('atualizacaoMapa.xlsx')

# Filtro para manter apenas os dados do ano atual
current_year = datetime.now().year
df_update = df_update[df_update['request'].dt.year == current_year]

# Concatenar o DataFrame atualizado com o DataFrame mestre filtrado
df_master = pd.concat([df_master, df_update], ignore_index=True)

# Atualizar a planilha 'Mapa Y24'
planilha_aberta = load_workbook('atualizacaoMapa.xlsx')
sheet_selecionada = planilha_aberta['Mapa Y23']

# Limpar a planilha 'Mapa Y24' antes de inserir os novos dados
sheet_selecionada.delete_rows(2, sheet_selecionada.max_row)

# Inserir os dados atualizados na planilha 'Mapa Y24'
for index, row in df_master.iterrows():
    sheet_selecionada.append(row.tolist())

# Salvar o arquivo com data de atualização no nome
data_atualizacao = datetime.now().strftime('%Y-%m-%d')
nome_arquivo = f'atualizacaoMapa_{data_atualizacao}.xlsx'
planilha_aberta.save(nome_arquivo)