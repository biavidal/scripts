from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def ajustar_largura_colunas_excel(excel_file, sheet_name):
    # Abre o arquivo Excel
    wb = load_workbook(excel_file)
    
    # Obtém a planilha
    ws = wb[sheet_name]
    if sheet_name=='Resumo':
        # Adiciona a string na primeira célula
        cell = ws.cell(row=1, column=1, value="eu amo o brasil, eu amo o brasil!!!")
        # Aplica formatação em negrito e centralizada
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Ajusta automaticamente a largura das colunas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(col[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salva o arquivo Excel com as colunas ajustadas
    wb.save(excel_file)
dados_mean = {"Data" : ['aaaaa', 'aaaaa', 'aaaaaa'],
                  "": ['bbbbb', 'bbbbbb', 'bbbbbb'],
                  "January": ['cccccc', 'cccccc', 'cccccc'],
                  "February": ['ddddd', 'ddddd', 'ddddd'],
                  "March": ['eeeeee', 'eeeeee', 'eeeee']}

dados_rate = {"Tabela" : ['aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa', 'aaaaaaaaaaaaaaaaaaaaaa'],
              "Rate": ['bbbbb', 'bbbbbb', 'bbbbbb', 'bbbbbbbb', 'bbbbbbb', 'bbbbbbb'],
              "January": ['cccccc', 'cccccc', 'cccccc', 'cccccccc', 'ccccccc', 'cccccccc'],
              "February": ['dddddddd', 'ddddd', 'ddddd', 'dddddddd', 'dddddddd', 'dddddddd'],
              "March": ['eeeeee', 'eeeeee', 'eeeee', 'eeeeee', 'eeeeee', 'eeeeee']}

df_mean = pd.DataFrame(dados_mean)
df_rate = pd.DataFrame(dados_rate)
# Imprime o DataFrame resultante
with pd.ExcelWriter('dados_usd.xlsx', engine='xlsxwriter') as writer:
    df_mean.to_excel(writer, index=False, startrow=2, sheet_name='Copper')
    df_rate.to_excel(writer, index=False, startrow=7, sheet_name='Resumo')

# Exemplo de uso da função
ajustar_largura_colunas_excel('dados_usd.xlsx', 'Resumo')
