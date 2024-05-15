from data_handler import load_data
from openpyxl import load_workbook
from datetime import datetime

def update_MapaFile(file_path, df_master, sheet_name='Mapa Y24'):
    planilha_aberta = load_workbook(file_path)
    sheet_selecionada = planilha_aberta[sheet_name]
    sheet_selecionada.delete_rows(2, sheet_selecionada.max_row)
    for index, row in df_master.iterrows():
        sheet_selecionada.append(row.tolist())
    data_atualizacao = datetime.now().strftime('%Y-%m-%d')
    nome_arquivo = f'atualizacaoMapa_{data_atualizacao}.xlsx'
    planilha_aberta.save(nome_arquivo)

# Carregar dados usando a função load_data
file_path, df_master = load_data('atualizacaoMapa.xlsx')

# Atualizar a planilha 'Mapa Y24'
update_MapaFile(file_path, df_master)