import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def stylize_and_merge_excel(df_copper, df_aluminium, output_file):
    # Criar um novo arquivo Excel
    workbook = Workbook()
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    writer.book = workbook

    # Estilizar e salvar a primeira planilha
    stylize_sheet(workbook, df_copper, 'Copper')

    # Estilizar e salvar a segunda planilha
    stylize_sheet(workbook, df_aluminium, 'Aluminium')

    # Salvar o arquivo Excel
    writer.save()

def stylize_sheet(workbook, df, sheet_name):
    # Adicionar a planilha ao arquivo Excel
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    worksheet = workbook[sheet_name]

    # Aplicar estilos diretamente às linhas
    for row_index, row in enumerate(worksheet.iter_rows(), start=1):
        if row_index == 1 or row_index == len(df) + 1:
            # Estilizar a primeira e última linhas
            border_style = 'thick' if row_index == 1 else 'thin'
            font = Font(bold=True)
        else:
            # Estilizar as outras linhas
            border_style = 'thin'
            font = None

        for col_index, cell in enumerate(row, start=1):
            # Verificar se a coluna contém 'metal' em alguma parte do nome
            col_name = worksheet.cell(row=1, column=col_index).value
            if col_name and 'metal' in col_name.lower():
                # Se a coluna contém 'metal', aplicar bordas apenas nas laterais esquerda e direita
                cell.border = Border(left=Side(style=border_style, color='000000'),
                                     right=Side(style=border_style, color='000000'))

            # Aplicar estilo negrito para a primeira e última linhas
            if font:
                cell.font = font

    # Centralizar o conteúdo das células
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Exemplo de uso
df_copper, df_aluminium = extract_excel(excel_files)  # Supondo que extract_excel seja uma função definida em outro lugar

stylize_and_merge_excel(df_copper, df_aluminium, 'styled_dataframe.xlsx')